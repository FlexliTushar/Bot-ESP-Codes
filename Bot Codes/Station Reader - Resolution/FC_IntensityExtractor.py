#!/usr/bin/env python3
"""
station_parser.py

Parses logs for:
  - FC: <n> | Intensity: { a, b, c }
  - Read time: <ms>
  - Current station: <ID>

For each Read time it groups the preceding FC+Intensity matches (since last Read time)
and associates them with the next Current station found after that Read time.

Output: station_wise.xlsx (one sheet per station). Each round is written as a vertical
block; rounds are placed side-by-side across columns.
"""
import re
import argparse
from openpyxl import Workbook
import sys

# -----------------------
# Regex (case-insensitive)
# -----------------------
FC_RE = re.compile(r"FC:\s*(\d+)\s*\|\s*Intensity:\s*{\s*(\d+)\s*,\s*(\d+)\s*,\s*(\d+)\s*}", re.IGNORECASE)
READTIME_RE = re.compile(r"Read time:\s*(\d+)", re.IGNORECASE)
STATION_RE = re.compile(r"Current station:\s*([A-Za-z0-9_]+)", re.IGNORECASE)

# sanitize sheet name to be Excel-safe (<=31 chars and allowed chars)
def sanitize_sheet_name(name: str) -> str:
    name = re.sub(r"[\\/*?:\[\]]", "_", name)
    return name[:31]

def parse_logs(text: str):
    fc_matches = list(FC_RE.finditer(text))
    read_matches = list(READTIME_RE.finditer(text))
    station_matches = list(STATION_RE.finditer(text))

    if not read_matches:
        return {}

    # Build indexable lists of starts for speed
    fc_positions = [(m.start(), m) for m in fc_matches]
    station_positions = [(m.start(), m) for m in station_matches]

    station_rounds = {}  # station_id -> list of rounds; round = {'read_time':int, 'columns':[{'fc':..., 'ints':[...]}]}

    for i, read_m in enumerate(read_matches):
        read_pos = read_m.start()
        prev_read_pos = read_matches[i-1].start() if i > 0 else -1
        next_read_pos = read_matches[i+1].start() if i+1 < len(read_matches) else None

        # FC matches that are after prev_read_pos and before this read_pos
        associated_fcs = [m for (p, m) in fc_positions if p > prev_read_pos and p < read_pos]

        # find station: first station match whose start > read_pos and (if next_read_pos exists) start < next_read_pos
        station_m = None
        for (p, m) in station_positions:
            if p > read_pos and (next_read_pos is None or p < next_read_pos):
                station_m = m
                break
        # if not found in that window, try the first station after read_pos at all
        if station_m is None:
            for (p, m) in station_positions:
                if p > read_pos:
                    station_m = m
                    break

        if station_m is None:
            continue

        # Try Current station first
        station_match = re.search(r"Current\s+station\s*:\s*(.*?)\s*\|", text[read_pos:next_read_pos] if next_read_pos else text[read_pos:])
        invalid_match = re.search(r"Invalid station code\s*:\s*(\S+)", text)

        # Determine search window for this read time
        search_text = text[read_pos:next_read_pos] if next_read_pos else text[read_pos:]

        # Try Current station first
        station_m_local = re.search(r"Current\s+station\s*:\s*([A-Za-z0-9_]+)", search_text, re.IGNORECASE)
        invalid_m_local = re.search(r"Invalid\s+station\s+code\s*[:\-]?\s*(\S+)", search_text, re.IGNORECASE)

        if station_m_local:
            station_id = station_m_local.group(1)
        elif invalid_m_local:
            station_id = invalid_m_local.group(1)
        else:
            station_id = "UnknownStation"

        read_time_val = int(read_m.group(1))

        # Build columns
        columns = []
        for m in associated_fcs:
            fc_val = int(m.group(1))
            ints = [int(m.group(2)), int(m.group(3)), int(m.group(4))]
            columns.append({'fc': fc_val, 'ints': ints})

        # If no columns found for this read time, still create a round with empty columns? skip to keep sheet readable
        if len(columns) == 0:
            # skip empty rounds
            continue

        station_rounds.setdefault(station_id, []).append({
            'read_time': read_time_val,
            'columns': columns
        })

    return station_rounds

def write_excel(station_rounds: dict, out_path: str):
    wb = Workbook()
    default_sheet = wb.active
    default_sheet.title = "Temp"

    for station_id, rounds in station_rounds.items():
        sheet_name = sanitize_sheet_name(station_id)
        ws = wb.create_sheet(title=sheet_name)

        start_col = 1
        for r_index, rnd in enumerate(rounds, start=1):
            col = start_col

            # Header
            ws.cell(row=1, column=col, value=f"Round {r_index}")
            ws.cell(row=2, column=col, value="Read Time (ms)")
            ws.cell(row=2, column=col+1, value=rnd['read_time'])

            row_ptr = 4
            for c_idx, colinfo in enumerate(rnd['columns'], start=1):
                ws.cell(row=row_ptr, column=col, value=f"Column {c_idx}")
                row_ptr += 1
                ws.cell(row=row_ptr, column=col, value="FC")
                ws.cell(row=row_ptr, column=col+1, value=colinfo['fc'])
                row_ptr += 1
                ws.cell(row=row_ptr, column=col, value="Intensity")
                # intensities occupy next 3 columns
                ws.cell(row=row_ptr, column=col+1, value=colinfo['ints'][0])
                ws.cell(row=row_ptr, column=col+2, value=colinfo['ints'][1])
                ws.cell(row=row_ptr, column=col+3, value=colinfo['ints'][2])
                # move pointer down with a blank row between columns
                row_ptr += 2

            # Move start_col right for next round (block width = 4 columns; leave 2-column gap)
            start_col += 6

    # Keep Temp only if no real sheets created
    if len(wb.sheetnames) > 1:
        wb.remove(default_sheet)

    try:
        wb.save(out_path)
    except Exception as e:
        print("Failed to save Excel file:", e)
        raise

def main():
    parser = argparse.ArgumentParser(description="Parse bot logs into station-wise Excel.")
    parser.add_argument("logfile", nargs="?", default="B11_d425_9Oct.txt", help="Path to input log file (default: logs.txt)")
    parser.add_argument("-o", "--out", default="B11_d425_9Oct_FC&Intensity_station_wise.xlsx", help="Output Excel file (default: station_wise.xlsx)")
    args = parser.parse_args()

    try:
        with open(args.logfile, "r", encoding="utf-8", errors="ignore") as f:
            txt = f.read()
    except FileNotFoundError:
        print("Log file not found:", args.logfile)
        sys.exit(1)

    station_rounds = parse_logs(txt)

    if not station_rounds:
        print("No station rounds parsed. Check that your log contains 'FC: ... | Intensity: {..}' and 'Read time:' and 'Current station:'.")
        # still write an empty workbook so openpyxl won't error
        wb = Workbook()
        wb.active.title = "NoData"
        wb.save(args.out)
        print("Empty workbook written to", args.out)
        return

    # Info
    for s, rounds in station_rounds.items():
        print(f"Station '{s}': {len(rounds)} rounds (columns per round: {[len(r['columns']) for r in rounds]})")

    write_excel(station_rounds, args.out)
    print("Excel written to", args.out)

if __name__ == "__main__":
    main()
