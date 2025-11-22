import re
import os
import openpyxl
from openpyxl.styles import PatternFill

# --- CONFIG ---
input_file = input("Enter log filename: ").strip()
base, ext = os.path.splitext(input_file)
output_excel = f"{base}-analyzed.xlsx"

# --- READ LOG FILE ---
with open(input_file, "r") as f:
    log_data = f.read()

# --- REGEX ---
station_pattern = re.compile(r"(Current station|Invalid station code)[:\-]?\s*(\S+)")
instcol_pattern = re.compile(
    r"\|\s*Inst\. Col:\s*(\d+)\s*\|\s*Inst\. FC:\s*(\d+)\s*\|\s*Inst\. Max\. Intensity:\s*\{\s*(.*?)\s*\}\s*\|\s*Inst\. Min\. Intensity:\s*\{\s*(.*?)\s*\}", re.DOTALL
)
intercol_pattern = re.compile(r"Col No\.\:\s*(\d+)\s*\|\s*Inter\. Col:\s*(\d+)\s*\|\s*Inter\. FC:\s*(\d+)")
pd_pattern = re.compile(r"Photodiode\s+(\d+)\s*:\s*([A-Za-z]+)")

# --- CREATE EXCEL ---
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Analysis"
ws.append(["Station", "Inter. Col", "Health", "Remarks"])

# Color fills
fill_healthy = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
fill_warning = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
fill_critical = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")

# --- FUNCTION ---
def process_station_logs(log):
    station_headers = list(station_pattern.finditer(log))
    station_ranges = []

    prev_end = 0
    for sh in station_headers:
        station_ranges.append((prev_end, sh.start(), sh.group(2)))
        prev_end = sh.end()
    if station_headers:
        station_ranges.append((prev_end, len(log), station_headers[-1].group(2)))

    for start, end, station_id in station_ranges:
        station_log = log[start:end]

        insts = []
        for m in instcol_pattern.finditer(station_log):
            insts.append({
                "start": m.start(),
                "end": m.end(),
                "col": m.group(1),
                "fc": m.group(2),
                "max": m.group(3).strip(),
                "min": m.group(4).strip()
            })

        inters = []
        for m in intercol_pattern.finditer(station_log):
            inters.append({
                "start": m.start(),
                "col": m.group(2),
                "fc": m.group(3),
                "col_no": m.group(1)
            })

        # --- collect rows for this station ---
        station_rows = []

        for inter in inters:
            matching_inst = None
            for inst in reversed([i for i in insts if i["start"] < inter["start"]]):
                if inst["col"] == inter["col"] and inst["fc"] == inter["fc"]:
                    matching_inst = inst
                    break
            if not matching_inst:
                continue

            later_insts = [i for i in insts if i["start"] > matching_inst["start"]]
            next_inst_start = later_insts[0]["start"] if later_insts else len(station_log)
            segment = station_log[matching_inst["end"]: next_inst_start]
            pd_matches = pd_pattern.findall(segment)

            # --- Health Logic (identical PD handling) ---
            health = "Healthy"
            for _, status in pd_matches:
                if status.lower() == "critical":
                    health = "Critical"
                    break
                elif status.lower() == "warning" and health != "Critical":
                    health = "Warning"

            remarks = f"Inst.Max: {matching_inst['max']} | Inst.Min: {matching_inst['min']}"
            if pd_matches:
                pd_status_str = ", ".join(f"PD{idx}:{status}" for idx, status in pd_matches)
                remarks += f" | {pd_status_str}"

            # collect row instead of writing now
            station_rows.append((station_id, inter["col"], health, remarks))

        # --- after loop: only write if at least one warning or critical ---
        if any(health in ("Warning", "Critical") for _, _, health, _ in station_rows):
            for row in station_rows:
                ws.append(row)
                fill = (
                    fill_critical if row[2] == "Critical"
                    else fill_warning if row[2] == "Warning"
                    else fill_healthy
                )
                ws[f"C{ws.max_row}"].fill = fill
        # After processing all Inter.Cols of this station → insert an empty separator row
        ws.append([])

# --- RUN ---
process_station_logs(log_data)
wb.save(output_excel)
print(f"Analysis saved to: {output_excel}")