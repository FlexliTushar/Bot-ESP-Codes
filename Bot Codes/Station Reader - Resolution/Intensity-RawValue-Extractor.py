import re
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

# === INPUT / OUTPUT FILE PATHS ===
input_file = "D16-B10-logs.txt"   # Replace with your actual log file name or path
output_file = "D16-B10-logs.xlsx"

# === READ THE LOG FILE ===
with open(input_file, "r", encoding="utf-8", errors="ignore") as f:
    log_data = f.read()

# === REGEX PATTERNS ===
# This pattern captures both valid stations (like I11, D4) and empty station cases
station_block_pattern = r"Current station:\s*([A-Z]?\d*)(.*?)(?=Current station:|$)"
reading_pattern = r"(\d+):\s*([\d]+),\s*([\d]+),\s*([\d]+)"

# === DICTIONARY: station → list of dataframes ===
stations_data = {}

# === PARSE EACH STATION BLOCK ===
for match in re.finditer(station_block_pattern, log_data, re.DOTALL):
    station_name = match.group(1).strip()
    block_content = match.group(2)

    # Handle invalid station codes if no station name present
    if not station_name:
        invalid_match = re.search(r"Invalid station code:\s*(\S+)", block_content)
        if invalid_match:
            station_name = f"Invalid station code {invalid_match.group(1)}"
        else:
            station_name = "Unknown_Station"

    # Extract readings
    readings = re.findall(reading_pattern, block_content)
    if readings:
        df = pd.DataFrame(readings, columns=["Index", "PD1", "PD2", "PD3"]).astype(int)
        stations_data.setdefault(station_name, []).append(df)

# === WRITE TO EXCEL (EACH SHEET PER STATION) ===
with pd.ExcelWriter(output_file, engine="openpyxl") as writer:
    # Sort sheet names alphabetically (optional but tidy)
    for station in sorted(stations_data.keys(), key=lambda x: x.lower()):
        rounds = stations_data[station]

        # Combine all rounds with a blank column in between
        spaced_rounds = []
        for i, df in enumerate(rounds):
            spaced_rounds.append(df)
            if i < len(rounds) - 1:
                spaced_rounds.append(pd.DataFrame({"": [""] * len(df)}))  # Gap column

        combined_df = pd.concat(spaced_rounds, axis=1)

        # Create headers with blank column headers for gaps
        headers = []
        for i, df in enumerate(rounds):
            headers += [f"Round {i+1} - Index", f"Round {i+1} - PD1", f"Round {i+1} - PD2", f"Round {i+1} - PD3"]
            if i < len(rounds) - 1:
                headers.append("")  # Blank column for gap
        combined_df.columns = headers

        combined_df.to_excel(writer, sheet_name=station[:31], index=False)  # Excel max sheet name length = 31

# === AUTO-RESIZE COLUMNS ===
wb = load_workbook(output_file)
for ws in wb.worksheets:
    for col in ws.columns:
        max_length = 0
        column = get_column_letter(col[0].column)
        for cell in col:
            try:
                val_len = len(str(cell.value))
                if val_len > max_length:
                    max_length = val_len
            except:
                pass
        ws.column_dimensions[column].width = max_length + 2

wb.save(output_file)
print(f"✅ Excel file created successfully: {output_file}")